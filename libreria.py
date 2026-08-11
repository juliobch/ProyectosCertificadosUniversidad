# libreria.py — COCINA: lectura, cruce y cálculo de certificados.
# Esta biblioteca contiene TODA la lógica de negocio.
# Las plantillas y rutas solo la llaman; aquí no se muestra nada al usuario.

# LADRILLO 1: LIBRERÍAS (módulos importados)
import unicodedata  # para quitar acentos al normalizar texto
from pathlib import Path  # para construir rutas que funcionen desde cualquier carpeta
from openpyxl import load_workbook  # lee archivos .xlsx (nunca los modifica)

# Ruta base del proyecto: la carpeta donde está ESTE archivo (libreria.py).
# Así la app encuentra Insumos/ sin importar desde dónde se ejecute.
CARPETA_PROYECTO = Path(__file__).resolve().parent


# ---------------------------------------------------------------------------
# LADRILLO 2: FUNCIÓN de normalización
# ---------------------------------------------------------------------------
def normalizar(texto):
    """Limpia un texto: quita acentos, pasa a minúsculas y quita espacios dobles.

    Esto es un LADRILLO importante: hace que 'Técnico en IA' y 'tecnico en ia'
    se vean iguales al momento de cruzar los dos archivos.
    """
    if texto is None:
        return ""  # VALOR por defecto si la celda está vacía
    # QUITA ACENTOS: descompone la letra (ej: é -> e + tilde) y borra la tilde
    sin_acentos = unicodedata.normalize("NFD", str(texto))
    sin_acentos = "".join(c for c in sin_acentos if unicodedata.category(c) != "Mn")
    # PASA A MINÚSCULAS y une espacios múltiples en uno solo
    return " ".join(sin_acentos.lower().split())


# ---------------------------------------------------------------------------
# LADRILLO 3: FUNCIÓN leer_maestro
# ---------------------------------------------------------------------------
def leer_maestro(ruta=None):
    """Lee el archivo maestro y devuelve una LISTA de DICCIONARIOS por estudiante.

    Cada estudiante es un dict:
        {'identificacion': int, 'nombre': str, 'correo': str,
         'programa': str, 'programa_norm': str, 'cohorte': str}

    Si no se pasa ruta, se usa la de Insumos/ junto al proyecto.
    """
    if ruta is None:
        # Ruta absoluta al Excel, calculada desde la carpeta del proyecto
        ruta = CARPETA_PROYECTO / "Insumos" / "Maestro_Estudiantes.xlsx"
    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active  # primera (y única) hoja del Excel

    estudiantes = []  # VARIABLE de tipo LISTA que guardará el resultado
    fila = 0  # VARIABLE CONTADORA: sirve para saltarnos la fila de encabezados

    # LADRILLO 4: BUCLE (for) que recorre todas las filas de la hoja
    for fila_datos in hoja.iter_rows(values_only=True):
        fila += 1
        if fila == 1:
            continue  # CONDICIONAL: la primera fila son los nombres de las columnas
        if fila_datos[0] is None:
            continue  # CONDICIONAL: fila vacía, la ignoramos

        # CONVERTIMOS TIPOS: la cédula viene como texto y la pasamos a entero (int)
        identificacion = int(fila_datos[0])
        nombre = str(fila_datos[1])
        correo = str(fila_datos[2])
        programa = str(fila_datos[3])
        cohorte = str(fila_datos[4])

        # Guardamos un DICCIONARIO por estudiante dentro de la lista
        estudiantes.append({
            "identificacion": identificacion,
            "nombre": nombre,
            "correo": correo,
            "programa": programa,
            "programa_norm": normalizar(programa),  # clave limpia para cruzar
            "cohorte": cohorte,
        })

    libro.close()  # cerramos el archivo (solo lectura)
    return estudiantes


# ---------------------------------------------------------------------------
# LADRILLO 5: FUNCIÓN leer_evaluaciones
# ---------------------------------------------------------------------------
def leer_evaluaciones(ruta=None):
    """Lee las notas/asistencia y devuelve una LISTA de DICCIONARIOS por fila.

    Cada registro es un dict:
        {'identificacion': int, 'programa': str, 'programa_norm': str,
         'modulo': str, 'nota': float, 'asistencia': float, 'fecha': str}

    Si no se pasa ruta, se usa la de Insumos/ junto al proyecto.
    """
    if ruta is None:
        # Ruta absoluta al Excel, calculada desde la carpeta del proyecto
        ruta = CARPETA_PROYECTO / "Insumos" / "Registro_Evaluaciones.xlsx"
    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active

    registros = []  # VARIABLE de tipo LISTA con todas las filas de notas
    fila = 0  # VARIABLE CONTADORA para saltar el encabezado

    # LADRILLO 6: BUCLE (for) sobre las filas del Excel de notas
    for fila_datos in hoja.iter_rows(values_only=True):
        fila += 1
        if fila == 1:
            continue  # encabezados
        if fila_datos[0] is None:
            continue  # fila vacía

        # CONVERSIÓN DE TIPOS: cédula a int, nota y asistencia a float (decimal)
        registros.append({
            "identificacion": int(fila_datos[0]),
            "programa": str(fila_datos[1]),
            "programa_norm": normalizar(fila_datos[1]),
            "modulo": str(fila_datos[2]),
            "nota": float(fila_datos[3]),
            "asistencia": float(fila_datos[4]),
            "fecha": str(fila_datos[5]),
        })

    libro.close()
    return registros


# ---------------------------------------------------------------------------
# LADRILLO 7: FUNCIÓN agrupar_evaluaciones
# ---------------------------------------------------------------------------
def agrupar_evaluaciones(registros):
    """Agrupa las notas por (identificacion, programa).

    Devuelve un DICCIONARIO cuya clave es (int, str_normalizada) y cuyo
    valor es una LISTA de dicts de notas de ese estudiante en ese programa.

    Esto implementa la regla: 'Se agrupa por Identificacion + Programa'.
    """
    grupos = {}  # VARIABLE de tipo DICCIONARIO

    # LADRILLO 8: BUCLE que recorre todos los registros de notas
    for r in registros:
        clave = (r["identificacion"], r["programa_norm"])  # TUPLA como llave
        if clave not in grupos:
            grupos[clave] = []  # si es la primera nota, creamos la lista
        grupos[clave].append(r)  # agregamos esta nota al grupo

    return grupos


# ---------------------------------------------------------------------------
# LADRILLO 9: FUNCIÓN calcular_promedio
# ---------------------------------------------------------------------------
def calcular_promedio(notas):
    """Recibe una LISTA de notas (números) y devuelve su promedio (float).

    Regla de AGENTS.md: Promedio = suma de Notas / cantidad de módulos cursados.
    """
    if not notas:  # CONDICIONAL: lista vacía -> no hay promedio
        return 0.0
    suma = 0.0  # VARIABLE ACUMULADORA: irá sumando las notas
    # LADRILLO 10: BUCLE (for) que suma nota por nota
    for n in notas:
        suma += n
    # DIVISIÓN: suma total / cantidad de módulos (len devuelve cuántos hay)
    return suma / len(notas)


# ---------------------------------------------------------------------------
# LADRILLO 11: FUNCIÓN decidir_certificado
# ---------------------------------------------------------------------------
def decidir_certificado(promedio, asistencia):
    """Aplica las reglas de AGENTS.md y devuelve el TIPO de certificado.

    Los límites INCLUYEN el valor (>= significa 'igual o mayor que').
    Valores posibles (TEXTOS constantes):
        'APROBACION', 'PARTICIPACION', 'SIN_CERTIFICADO'
    """
    # LADRILLO 12: CONDICIONALES encadenados (si / si no si / si no)
    if promedio >= 70 and asistencia >= 80:
        # AND: se deben cumplir las DOS condiciones juntas
        return "APROBACION"
    elif promedio < 70 and asistencia >= 80:
        return "PARTICIPACION"
    else:
        # Cualquier otro caso: asistencia < 80, o sin datos -> sin certificado
        return "SIN_CERTIFICADO"


# ---------------------------------------------------------------------------
# LADRILLO 13: FUNCIÓN calcular_resultados
# ---------------------------------------------------------------------------
def calcular_resultados(estudiantes, registros):
    """Cruza maestro + evaluaciones y devuelve la lista de resultados finales.

    Cada resultado es un dict:
        {'identificacion', 'nombre', 'programa', 'cohorte',
         'modulos', 'promedio', 'asistencia', 'tipo'}
    Además devuelve aparte las INCONSISTENCIAS del cruce:
        estudiantes sin notas y evaluaciones sin estudiante en el maestro.
    """
    grupos = agrupar_evaluaciones(registros)  # llamada a la FUNCIÓN del ladrillo 7

    resultados = []  # lista final que se mostrará en la tabla
    sin_notas = []  # INCONSISTENCIA: estudiantes del maestro SIN evaluaciones
    sin_maestro = []  # INCONSISTENCIA: evaluaciones SIN estudiante en el maestro
    ids_validos = set()  # VARIABLE de tipo CONJUNTO: cédulas que SÍ están en el maestro

    # LADRILLO 14: BUCLE que recorre cada estudiante del maestro
    for e in estudiantes:
        ids_validos.add(e["identificacion"])  # recordamos qué cédulas son oficiales
        clave = (e["identificacion"], e["programa_norm"])

        if clave not in grupos:
            # CONDICIONAL: el estudiante NO cursó este programa -> sin datos
            sin_notas.append(e)
            resultados.append({
                "identificacion": e["identificacion"],
                "nombre": e["nombre"],
                "programa": e["programa"],
                "cohorte": e["cohorte"],
                "modulos": 0,
                "promedio": 0.0,
                "asistencia": 0.0,
                "tipo": "SIN_CERTIFICADO",
            })
            continue  # pasamos al siguiente estudiante

        # Tomamos todas las notas de ese (cédula, programa)
        notas_grupo = grupos[clave]
        notas = [n["nota"] for n in notas_grupo]          # LISTA de notas
        asistencias = [n["asistencia"] for n in notas_grupo]  # LISTA de asistencias

        # Llamamos a las FUNCIONES de los ladrillos 9 y 11
        promedio = calcular_promedio(notas)
        asistencia = calcular_promedio(asistencias)  # reutilizamos el mismo cálculo
        tipo = decidir_certificado(promedio, asistencia)

        resultados.append({
            "identificacion": e["identificacion"],
            "nombre": e["nombre"],
            "programa": e["programa"],
            "cohorte": e["cohorte"],
            "modulos": len(notas),  # cantidad de módulos cursados
            "promedio": promedio,
            "asistencia": asistencia,
            "tipo": tipo,
        })

    # LADRILLO 15: BUCLE para encontrar evaluaciones sin estudiante en el maestro
    for clave in grupos:
        if clave[0] not in ids_validos:
            # El (cédula, programa) NO existe en el maestro -> inconsistencia
            sin_maestro.append({
                "identificacion": clave[0],
                "programa": grupos[clave][0]["programa"],  # nombre legible
                "modulos": len(grupos[clave]),
            })

    return resultados, sin_notas, sin_maestro


# ---------------------------------------------------------------------------
# LADRILLO 16: FUNCIÓN resumen_datos
# ---------------------------------------------------------------------------
def resumen_datos(estudiantes, registros):
    """Devuelve un DICCIONARIO con el resumen pedido:
    estudiantes por programa, cantidad de evaluaciones, módulos existentes
    y rango de notas.
    """
    # Estudiantes por programa: DICCIONARIO que cuenta
    por_programa = {}  # clave: nombre del programa, valor: contador
    for e in estudiantes:
        nombre_programa = e["programa"]
        if nombre_programa not in por_programa:
            por_programa[nombre_programa] = 0  # CONDICIONAL: arrancamos el conteo
        por_programa[nombre_programa] += 1

    # Módulos existentes: lista sin repetidos (usamos un CONJUNTO)
    modulos = set()  # VARIABLE de tipo CONJUNTO (no admite duplicados)
    for r in registros:
        modulos.add(r["modulo"])  # BUCLE que llena el conjunto

    # Rango de notas: min y max (recorremos todas las notas)
    notas_min = None
    notas_max = None
    for r in registros:
        if notas_min is None or r["nota"] < notas_min:
            notas_min = r["nota"]  # CONDICIONAL: nueva nota mínima
        if notas_max is None or r["nota"] > notas_max:
            notas_max = r["nota"]  # CONDICIONAL: nueva nota máxima

    return {
        "total_estudiantes": len(estudiantes),
        "total_registros": len(registros),
        "por_programa": por_programa,
        "modulos": sorted(modulos),
        "notas_min": notas_min,
        "notas_max": notas_max,
    }
